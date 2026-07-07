"""微信 Excel 账单导入 → 按支付方式+币种匹配账户"""
INCOME_OK = {"已存入零钱", "已收钱"}
EXPENSE_OK = {"支付成功", "已转账", "对方已收钱"}

from ..models import FOREIGN_EXCHANGE_KEYWORDS


def _resolve_account(conn, payment_method: str, currency: str) -> int | None:
    acct = conn.execute(
        "SELECT id FROM accounts WHERE name=? AND currency=? AND is_active=1",
        (payment_method, currency),
    ).fetchone()
    if acct:
        return acct["id"]
    acct = conn.execute(
        "SELECT id FROM accounts WHERE name=? AND is_active=1 LIMIT 1",
        (payment_method,),
    ).fetchone()
    return acct["id"] if acct else None


def import_wechat(path: str):
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要 openpyxl: pip install openpyxl")
        return

    from ..db import get_db
    from ..txn import insert_txn

    conn = get_db()

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_row_i = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if row[0] and "交易时间" in str(row[0]):
            header_row_i = i
            break
    if not header_row_i:
        print("❌ 无法找到微信账单表头")
        conn.close()
        return

    header = [str(c or "") for c in next(ws.iter_rows(min_row=header_row_i, max_row=header_row_i, values_only=True))]
    h = {col: idx for idx, col in enumerate(header)}

    txns = []
    account_mismatch = 0

    for row in ws.iter_rows(min_row=header_row_i + 1, values_only=True):
        if not row or not any(v for v in row if v is not None):
            continue
        vals = [str(c or "") for c in row]

        direction = vals[h["收/支"]] if "收/支" in h else ""
        status = vals[h["当前状态"]] if "当前状态" in h else ""

        if direction == "支出" and status not in EXPENSE_OK:
            continue
        if direction == "收入":
            is_refund = "退款" in status
            if not is_refund and status not in INCOME_OK:
                continue

        try:
            amount = float(vals[h["金额(元)"]])
        except (ValueError, KeyError):
            continue

        if direction == "支出":
            amount = -amount
        elif direction == "收入":
            pass
        else:
            continue

        if amount == 0:
            continue

        payment_method = vals[h["支付方式"]] if "支付方式" in h else ""
        date_raw = vals[h["交易时间"]] if "交易时间" in h else ""
        date_str = date_raw[:10].replace("/", "-")
        counterparty = vals[h["交易对方"]] if "交易对方" in h else ""
        desc = vals[h["商品"]] if "商品" in h else ""

        # 币种 — 微信账单有"币种"列
        currency = "CNY"
        for cur_key in ("币种", "交易币种"):
            if cur_key in h:
                raw_cur = vals[h[cur_key]].strip().upper()
                if raw_cur in ("CNY", "USD", "HKD"):
                    currency = raw_cur
                break

        # 查找账户
        account_id = _resolve_account(conn, payment_method, currency)
        if account_id is None:
            print(f"  ⚠️ 未找到账户: 支付方式='{payment_method}' 币种={currency}，跳过该笔")
            account_mismatch += 1
            continue

        is_refund = "退款" in status
        is_exchange = any(kw in (desc + counterparty) for kw in FOREIGN_EXCHANGE_KEYWORDS)
        is_credit_repay = "信用卡还款" in (desc + counterparty)

        # 分类
        if is_exchange:
            category = "transfer"
        elif is_credit_repay and amount < 0:
            category = "transfer"
        elif is_refund and amount > 0:
            category = "expense"
        elif amount < 0:
            category = "expense"
        else:
            category = "income"

        txns.append({
            "date": date_str,
            "amount": amount,
            "account_id": account_id,
            "category": category,
            "counterparty": counterparty,
            "description": desc or counterparty,
            "payment_method": payment_method,
            "source_bill": "wechat",
            "source_file": path,
        })

    # 去重插入
    dedup_skipped = 0
    new_count = 0
    inserted_keys = set()
    for t in txns:
        key = (t["date"], round(t["amount"], 2), t["category"],
               t["counterparty"], t["description"], t["payment_method"])
        if key in inserted_keys:
            dedup_skipped += 1
            continue
        existing = conn.execute(
            """SELECT 1 FROM transactions WHERE date=? AND amount=?
               AND source_bill='wechat' AND category=? AND counterparty=?
               AND description=? AND payment_method=? LIMIT 1""",
            (t["date"], t["amount"], t["category"], t["counterparty"],
             t["description"], t["payment_method"]),
        ).fetchone()
        if existing:
            dedup_skipped += 1
            continue
        inserted_keys.add(key)
        insert_txn(conn, **t)
        new_count += 1

    total_skip = account_mismatch + dedup_skipped
    conn.execute(
        "INSERT INTO import_log(source_bill, filename, total, new, skipped) VALUES (?, ?, ?, ?, ?)",
        ("wechat", path, len(txns), new_count, total_skip),
    )
    conn.commit()
    conn.close()

    print(f"✅ 微信导入完成: 新增{new_count}条, 跳过{total_skip}条")
    if account_mismatch:
        print(f"   ⚠️ 其中 {account_mismatch} 条因找不到匹配账户跳过")
