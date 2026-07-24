"""convert — 账单 → 统一CSV"""
import hashlib
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import os
import tempfile


# 消费平台推断规则 — 从交易对方/描述中识别
PLATFORM_RULES = [
    ({"滴滴", "嘀嘀", "广州骑安", "先乘车后付款"}, "滴滴"),
    ({"高德打车", "高德地图", "高德信息技术", "高德云信"}, "高德"),
    ({"麦当劳"}, "麦当劳"),
    ({"汉堡王"}, "汉堡王"),
    ({"肯德基", "KFC"}, "肯德基"),
    ({"必胜客"}, "必胜客"),
    ({"海底捞"}, "海底捞"),
    ({"食其家"}, "食其家"),
    ({"争鲜寿司"}, "争鲜寿司"),
    ({"和府捞面"}, "和府捞面"),
    ({"袁记云饺"}, "袁记云饺"),
    ({"霸王茶姬"}, "霸王茶姬"),
    ({"茶百道"}, "茶百道"),
    ({"瑞幸咖啡", "luckin", "瑞幸", "luckincoffee"}, "瑞幸咖啡"),
    ({"奈雪"}, "奈雪"),
    ({"星巴克", "starbucks"}, "星巴克"),
    ({"LINLEE", "林里"}, "LINLEE林里"),
    ({"便利蜂", "梦想蜂"}, "便利蜂"),
    ({"7-11", "7-ELEVEN"}, "7-11"),
    ({"喜家德"}, "喜家德"),
    ({"新又好", "NewUhoo"}, "新又好"),
    ({"西部马华"}, "西部马华"),
    ({"立普世"}, "立普世"),
    ({"小红书"}, "小红书"),
    ({"bilibili", "哔哩哔哩", "B站", "b站"}, "B站"),
    ({"DeepSeek", "深度求索"}, "DeepSeek"),
    ({"携程", "去哪儿"}, "携程"),
    ({"飞猪"}, "飞猪"),
    ({"首开易生活", "预付费充电"}, "首开易生活"),
    ({"哈啰好物", "上海钧哈", "哈啰", "电动车租车"}, "哈啰"),
    ({"美团平台商户", "新渔阳滑雪场"}, "美团"),
    ({"先骑后付"}, "美团"),
    ({"北京象鲜科技", "小象超市", "美团买菜"}, "美团"),
    ({"三快在线", "三快科技"}, "美团"),
    ({"微信红包", "群收款", "转账"}, "微信"),
    ({"猫眼"}, "猫眼"),
    ({"Steam"}, "Steam"),
    ({"PIXIV", "FANBOX", "pixiv"}, "pixiv"),
    ({"中国电信"}, "中国电信"),
    ({"杭州乐读", "网易云音乐"}, "网易云音乐"),
    ({"天猫"}, "天猫"),
    ({"淘宝"}, "淘宝"),
    ({"拼多多"}, "拼多多"),
    ({"京东", "京东超市"}, "京东"),
]



def _rec_date(rec: dict) -> str:
    return str(rec.get("date") or rec.get("occurred_at") or "")

def _infer_platform(counterparty: str, description: str, source: str) -> str:
    """从交易对方/说明推断消费平台，只识别公司级/连锁品牌，无匹配返回空"""
    text = (counterparty + " " + description).lower()
    # 已知误导排除：北京东子 含"京东"子串但无关
    text = text.replace("北京东子", "")

    for keywords, platform in PLATFORM_RULES:
        for kw in keywords:
            if kw.lower() in text:
                return platform
    return ""


# 支付源推断规则 — 信用卡账单中判断用了哪种支付
PAYMENT_SOURCE_RULES = [
    ("美团支付", "美团支付"),
    ("京东支付", "京东支付"),
    ("财付通(银联云闪付)", "银联云闪付"),
    ("财付通", "微信支付"),
    ("支付宝", "支付宝"),
    ("网银在线", "网银在线"),
    ("Apple.com/bill", "Apple Pay"),
    ("Apple Pay", "Apple Pay"),
    ("拼多多支付", "拼多多支付"),
    ("程支付", "携程"),
    ("抖音支付", "抖音支付"),
    ("银联云闪付", "银联云闪付"),
    ("云闪付", "云闪付"),
]


def _infer_payment_source(bill_type: str, counterparty: str, description: str) -> str:
    """推断支付源"""
    if bill_type == "ccb_debit":
        return "建行储蓄卡"
    if bill_type == "icbc_debit":
        return "银行卡"
    if bill_type == "alipay":
        return "支付宝"
    if bill_type == "wechat":
        return "微信"
    # ICBC 类：从描述/对方名中识别支付源
    text = (counterparty + " " + description).lower()
    for keyword, source in PAYMENT_SOURCE_RULES:
        if keyword.lower() in text:
            return source
    return "银行卡"


# counterparty 中去掉的支付源前缀（对应 PAYMENT_SOURCE_RULES 的关键词）
# 注意：Apple.com/bill 和 Apple Pay 不加前缀处理
STRIP_PREFIXES = [
    "美团支付-",
    "京东支付-",
    "财付通(银联云闪付)",
    "财付通-",
    "支付宝-",
    "网银在线-",
    "拼多多支付-",
    "程支付-",
    "抖音支付-",
]


def _strip_payment_prefix(counterparty: str) -> str:
    """去掉交易对方中已知的支付源前缀，保留纯商家名"""
    import re
    if not counterparty:
        return counterparty
    # 先去掉分期标记（如 "10/24 ", "5/12 "），避免破坏 startswith
    stripped = re.sub(r"^\d+/\d+\s*", "", counterparty)
    for prefix in STRIP_PREFIXES:
        if stripped.startswith(prefix):
            stripped = stripped[len(prefix):].lstrip("-")
            break
    return stripped


# O2O 平台前缀（从交易对方开头去掉，但结果非空才去掉）
O2O_PREFIXES = [
    "美团App",
    "饿了么",
    "大众点评",
    "高德团购",
]

# 品牌匹配后从 leftover 中去除的电商后缀
ECOMMERCE_SUFFIXES = [
    "自营旗舰店",
    "官方旗舰店",
    "旗舰店",
    "官方店",
    "专卖店",
    "专营店",
]


def _strip_platform_prefix(counterparty: str) -> str:
    """去掉交易对方中的O2O平台前缀（美团App/饿了么/大众点评/高德团购）。
    若去掉后为空，返回原值不变。"""
    if not counterparty:
        return counterparty
    for prefix in O2O_PREFIXES:
        if counterparty.startswith(prefix):
            stripped = counterparty[len(prefix):].lstrip()
            if stripped:
                return stripped
            return counterparty
    return counterparty


def _extract_leftover(cp: str, brand: str) -> str:
    """品牌匹配后，从原交易对方名中提取有用剩余文本。

    1. 依次去掉品牌关键词（优先最长）
    2. 去掉 O2O 平台前缀
    3. 去掉常见电商后缀（自营旗舰店等）
    4. 清理首尾标点/空格
    """
    import re

    if not cp:
        return ""

    # 找到该品牌对应的所有关键词
    brand_keywords = []
    for keywords, b in PLATFORM_RULES:
        if b == brand:
            brand_keywords = list(keywords)
            break

    if not brand_keywords:
        return ""

    leftover = cp
    cp_lower = leftover.lower()

    # 从长到短依次去掉关键词，避免被短关键词吃掉不该吃的部分
    for kw in sorted(brand_keywords, key=len, reverse=True):
        if kw.lower() in cp_lower:
            leftover = re.sub(re.escape(kw), "", leftover, flags=re.IGNORECASE)
            cp_lower = leftover.lower()

    # 去掉 O2O 平台前缀
    leftover = _strip_platform_prefix(leftover)

    # 去掉常见电商后缀
    for suffix in ECOMMERCE_SUFFIXES:
        if leftover.endswith(suffix):
            leftover = leftover[: -len(suffix)]
            break

    # 清理首尾标点/括号/空白
    leftover = re.sub(r"^[\s（(【\[、,，\-—]+", "", leftover)
    leftover = re.sub(r"[\s）)】\]、,，\-—]+$", "", leftover)

    return leftover.strip()


def _normalize_counterparty(raw_cp: str, raw_desc: str, source: str) -> tuple[str, str]:
    """三级回退规范化交易对方名。

    1. 去掉支付源前缀（_strip_payment_prefix）
    2. 品牌匹配（_infer_platform）→ 若命中，counterty=品牌名，剩余文本搬移到 description
       - 特殊：先骑后付 → cp=美团，desc=先骑后付
    3. O2O 平台前缀剥离（_strip_platform_prefix）→ 若变化，counterty=剥离结果
    4. 无匹配 → 原样返回
    """
    if not raw_cp:
        return (raw_cp, raw_desc)

    # Stage 1: 去掉支付源前缀
    cp = _strip_payment_prefix(raw_cp)
    desc = raw_desc

    # Stage 2: 品牌匹配
    brand = _infer_platform(cp, desc, source)
    if brand:
        # 特殊：先骑后付 → 美团
        if "先骑后付" in cp:
            new_desc = "先骑后付"
            if raw_desc:
                new_desc = f"先骑后付|{raw_desc}"
            return ("美团", new_desc)

        cp_before = cp
        cp = brand
        leftover = _extract_leftover(cp_before, brand)
        # 仅当原始描述为空或等于原始交易对方时，才用 leftover 替换
        if leftover and (not raw_desc or raw_desc == raw_cp):
            desc = leftover
        # 对于中介平台（淘宝/高德/美团等），尝试从描述中提取真实商户
        intermediary_brands = {"淘宝", "天猫", "美团", "高德", "大众点评"}
        if brand in intermediary_brands and raw_desc:
            import re
            if brand in ("淘宝", "天猫"):
                # 淘宝/天猫：description 首段含真实商户名
                # 提取到已知分隔符为止，无分隔符则整段作为商户名
                m = re.match(r"^(.+?)(?:外卖订单|订单|购物车)", raw_desc)
                if m:
                    merchant = m.group(1).strip(" ·-—")
                else:
                    merchant = raw_desc
                # 剔除已知平台相关行
                if merchant in ("超级吃货卡", "外卖红包", "淘宝"):
                    pass  # 保持 cp=淘宝
                elif merchant.startswith("淘宝"):
                    pass
                else:
                    # 剥离地址信息（括号内内容）
                    merchant = re.sub(r"[（(][^)）]*[)）]", "", merchant).strip()
                    if merchant and len(merchant) >= 2:
                        cp = merchant
                        desc = raw_desc
            elif brand == "高德":
                # 高德到店消费：description 含"XX商户 - 高德地图"模式
                m = re.match(r"^(.+?)\s*-\s*高德地图", raw_desc)
                if m:
                    cp = m.group(1).strip()
                    desc = raw_desc
                # 高德打车 → cp=高德 保持不变（正确）
        return (cp, desc)

    # Stage 3: O2O 平台前缀剥离
    stripped = _strip_platform_prefix(cp)
    if stripped != cp:
        return (stripped, desc)

    # No match
    return (cp, desc)


def _stable_short_hash(*parts: str, length: int = 12) -> str:
    normalized = [" ".join(str(part or "").split()) for part in parts]
    payload = "|".join(normalized)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:length]


REFUND_MATCH_MAX_DAYS = 30
REFUND_AUTO_PASS_MAX_DAYS = 14


def _counterparty_matches(exp_cpy: str, ref_cpy: str) -> bool:
    """判断交易双方是否指向同一实体 — 精确/子串/前缀匹配"""
    if exp_cpy == ref_cpy:
        return True
    # 双方都不为空时做子串匹配
    if exp_cpy and ref_cpy and len(exp_cpy) >= 2 and len(ref_cpy) >= 2:
        if exp_cpy in ref_cpy or ref_cpy in exp_cpy:
            return True
    return False


def _parse_record_datetime(value: str) -> datetime:
    fmt = "%Y-%m-%d" if len(value) == 10 else "%Y-%m-%d %H:%M:%S"
    return datetime.strptime(value, fmt)


def _specific_payment_account(value: str) -> str:
    generic_methods = {
        "银行卡",
        "支付宝",
        "微信",
        "微信支付",
        "美团支付",
        "京东支付",
        "网银在线",
        "Apple Pay",
        "拼多多支付",
        "携程",
        "抖音支付",
        "银联云闪付",
        "云闪付",
    }
    value = (value or "").strip()
    if not value or value in generic_methods:
        return ""
    return value


def _icbc_credit_offset_text(counterparty: str, description: str) -> str:
    return " ".join(
        part.strip() for part in (counterparty or "", description or "") if part and part.strip()
    )


def _classify_icbc_credit_offset_type(counterparty: str, description: str) -> str:
    text = _icbc_credit_offset_text(counterparty, description)
    if "减免年费" in text:
        return "fee_reversal"
    if "刷卡金入账" in text or "刷卡金退款" in text:
        return "benefit_rebate"
    if "返现" in text or "Rebate" in text:
        return "campaign_cashback"
    if "退货" in text or "退款" in text:
        return "merchant_refund"
    return ""


def _build_icbc_credit_offset_income(rec: dict, offset_type: str) -> dict:
    return {
        **rec,
        "offset_type": offset_type,
        "offset_strength": "strong",
        "offset_action": "keep_as_offset_income",
    }


def _icbc_credit_offset_cluster(value: str, description: str) -> str:
    text = f"{value or ''} {description or ''}"
    if "中国铁路网络有限公司" in text:
        return "railway_travel"
    if any(token in text for token in ("京东", "网银在线")):
        return "ecommerce_jd"
    if any(token in text for token in ("美团", "北京象鲜科技有限公司")):
        return "local_life_meituan"
    if "自助侠" in text:
        return "device_service"
    if any(token in text for token in ("携程", "去哪儿", "上海顺途科技有限公司")):
        return "rideshare_travel"
    if any(token in text for token in ("拼多多", "抖音")):
        return "group_buy_food"
    return ""


def _icbc_credit_account_cluster(payment_method: str, card_number: str) -> str:
    card_tail = (card_number or "").strip()
    if card_tail:
        return f"icbc_credit_card_{card_tail}"
    return f"icbc_credit_channel_{payment_method or 'unknown'}"


def _icbc_debit_refund_cluster(counterparty: str, description: str) -> str:
    text = f"{counterparty or ''} {description or ''}"
    if "支付宝" in text:
        return "alipay_refund"
    if any(token in text for token in ("财付通", "深圳市财付通支付科技有限公司")):
        return "tenpay_refund"
    if any(token in text for token in ("中国银联无卡快捷支付业务专户", "银联无卡支付业务")):
        return "unionpay_refund"
    if any(token in text for token in ("京东", "网银在线")):
        return "jd_refund"
    if "淘宝" in text:
        return "taobao_refund"
    return ""


def _icbc_debit_account_cluster(rec: dict) -> str:
    payment_method = (rec.get("payment_method", "") or "").strip()
    if payment_method:
        return f"icbc_debit_channel_{payment_method}"
    return "icbc_debit_channel_unknown"


def _ccb_refund_cluster(rec: dict) -> str:
    location = (rec.get("location", "") or "").strip()
    payment_method = (rec.get("payment_method", "") or "").strip()
    counterparty = (rec.get("counterparty", "") or "").strip()
    raw_cp = (rec.get("_raw_cp", "") or "").strip()
    location_cp = (rec.get("_ccb_location_cp", "") or "").strip()
    text = " ".join(part for part in (location, payment_method, counterparty, raw_cp, location_cp) if part)
    if location.startswith("财付通-") or "微信" in payment_method:
        return "ccb_wechat"
    if location.startswith("支付宝-") or "支付宝" in payment_method:
        return "ccb_alipay"
    if location.startswith("美团支付-") or "美团" in payment_method:
        return "ccb_meituan"
    if "PAYPAL" in text.upper() or "PayPal" in payment_method:
        return "ccb_paypal"
    if payment_method.startswith("建行储蓄卡"):
        return "ccb_card"
    return ""


def _refund_matches_basic_constraints(exp: dict, ref: dict, ref_amt: float, remaining_amount: float) -> bool:
    exp_account = _specific_payment_account(exp.get("payment_method", ""))
    ref_account = _specific_payment_account(ref.get("payment_method", ""))
    if exp_account and ref_account and exp_account != ref_account:
        return False
    if ref_amt > remaining_amount:
        return False
    if _rec_date(exp) > _rec_date(ref):
        return False
    delta_days = (_parse_record_datetime(_rec_date(ref)) - _parse_record_datetime(_rec_date(exp))).days
    if delta_days > REFUND_MATCH_MAX_DAYS:
        return False
    return True


def _refund_source_signal(ref: dict) -> str:
    return ref.get("_refund_signal", "")


def _is_icbc_credit_untrusted_merchant_text(value: str) -> bool:
    import re

    value = (value or "").strip()
    if not value:
        return True
    if value in {"A", "F", "D", "工商", "银行", "中国", "请扫描二维码"}:
        return True
    if re.fullmatch(r"[:\d\s.-]+", value):
        return True
    if len(value) <= 1:
        return True
    return False


def _refund_signal_is_strong(ref: dict) -> bool:
    return _refund_source_signal(ref) in {
        "alipay_status",
        "alipay_category_nocount",
        "wechat_status",
        "icbc_credit_return",
        "icbc_debit_refund",
        "ccb_debit_refund",
        "ccb_debit_desc",
    }


def _alipay_refund_signal(*, txn_type: str, txn_status: str, direction: str, description: str) -> str:
    if txn_status == "退款成功":
        return "alipay_status"
    if txn_type == "退款" and direction == "不计收支":
        return "alipay_category_nocount"
    if "退款" in description:
        return "alipay_desc"
    return ""


def _classify_refund_match(*, ref: dict, rule_hint: str, exact_amt: bool,
                           candidate_count: int, expense: dict) -> str:
    if rule_hint in {"refund_desc_fallback", "refund_gross_candidate"}:
        return "weak"
    if not _refund_signal_is_strong(ref):
        return "weak"
    if ref.get("_refund_signal") == "icbc_credit_return" and ref.get("offset_type") == "merchant_refund":
        trusted = ref.get("_icbc_refund_merchant_trusted", False)
        if not trusted:
            return "weak"
        if rule_hint not in {"refund_raw_cp_match", "refund_cp_match"}:
            return "weak"
        if not ref.get("_icbc_refund_same_cluster", False):
            return "weak"
        if not ref.get("_icbc_refund_same_account_cluster", False):
            return "weak"
    if ref.get("_refund_signal") == "icbc_debit_refund":
        if rule_hint not in {"refund_raw_cp_match", "refund_cp_match"}:
            return "weak"
        if candidate_count == 1:
            return "strong"
        if not ref.get("_icbc_debit_refund_same_cluster", False):
            return "weak"
        if not ref.get("_icbc_debit_refund_same_account_cluster", False):
            return "weak"
    if ref.get("_refund_signal") == "ccb_debit_refund":
        if rule_hint not in {"refund_raw_cp_match", "refund_cp_match"}:
            return "weak"
        if not ref.get("_ccb_refund_same_cluster", False):
            return "weak"
        if not exact_amt:
            return "weak"
        if candidate_count == 1:
            return "strong"
        return "weak"
    delta_days = (_parse_record_datetime(_rec_date(ref)) - _parse_record_datetime(_rec_date(expense))).days
    order_locked_hints = {
        "refund_merchant_order_match",
        "refund_txn_base_match",
        "refund_desc_order_match",
        "refund_wechat_meituan_order",
    }
    strong_signal = _refund_source_signal(ref)
    if delta_days > REFUND_AUTO_PASS_MAX_DAYS and rule_hint not in order_locked_hints:
        desc_confirms = _refund_desc_confirms_match(ref.get("note", ""), expense.get("note", ""))
        if not (
            candidate_count == 1
            and exact_amt
            and strong_signal == "alipay_status"
            and rule_hint == "refund_cp_match"
            and ref.get("payment_method", "") == expense.get("payment_method", "")
            and ref.get("counterparty", "") == expense.get("counterparty", "")
            and desc_confirms
        ):
            return "weak"
    if rule_hint == "refund_raw_cp_match" and "***" in expense.get("counterparty", ""):
        return "weak"
    return "strong"


def _build_refund_tracking_pair(*, expense: dict, refund: dict, match_type: str,
                               rule_hint: str, match_strength: str,
                               candidate_count: int) -> dict:
    return {
        "expense": dict(expense),
        "refund": dict(refund),
        "match_type": match_type,
        "rule_hint": rule_hint,
        "match_strength": match_strength,
        "candidate_count": candidate_count,
        "source_refund_signal": _refund_source_signal(refund),
    }


def _build_convert_fact_rows(rows: list[dict], tracking_pairs) -> list[dict]:
    combined: dict[str, dict] = {}

    def merge_row(rec: dict):
        fact_id = rec.get("_fact_id") or rec.get("record_id")
        if not fact_id:
            return
        if fact_id not in combined:
            combined[fact_id] = dict(rec)
            return
        existing = combined[fact_id]
        for key, value in rec.items():
            if key not in existing or existing.get(key, "") in {"", None}:
                existing[key] = value

    for row in rows:
        merge_row(row)
    for pair in tracking_pairs:
        merge_row(pair.get("expense", {}))
        merge_row(pair.get("refund", {}))

    fact_rows = list(combined.values())
    fact_rows.sort(key=lambda r: ((r.get("date") or r.get("occurred_at") or ""), r.get("amount", 0)))
    return fact_rows


def _attach_tracking_metadata(rows: list[dict], tracking_pairs) -> list[dict]:
    by_fact_id = {}
    for idx, row in enumerate(rows, 1):
        record_id = row.get("record_id") or row.get("_fact_id") or f"c_{idx:06d}"
        row["record_id"] = record_id
        row.setdefault("proposed_action", "leave_as_is")
        row.setdefault("offset_group", "")
        row.setdefault("offset_role", "")
        row.setdefault("offset_strength", "")
        row.setdefault("offset_source", "")
        row.setdefault("offset_rule_hint", "")
        row.setdefault("offset_match_type", "")
        by_fact_id[row.get("_fact_id", record_id)] = row

    expense_groups: dict[str, list[str]] = {}
    expense_strengths: dict[str, list[str]] = {}
    expense_sources: dict[str, list[str]] = {}
    expense_rules: dict[str, list[str]] = {}
    expense_match_types: dict[str, list[str]] = {}

    for group_index, pair in enumerate(tracking_pairs, 1):
        expense_id = pair["expense"].get("_fact_id")
        refund_id = pair["refund"].get("_fact_id")
        if not expense_id or not refund_id:
            continue
        expense_row = by_fact_id.get(expense_id)
        refund_row = by_fact_id.get(refund_id)
        if not expense_row or not refund_row:
            continue
        group = f"refund_{group_index:06d}"
        strength = pair.get("match_strength", "")
        source = pair.get("source_refund_signal", "")
        rule_hint = pair.get("rule_hint", "")
        match_type = pair.get("match_type", "")

        refund_row["offset_group"] = group
        refund_row["offset_role"] = "refund"
        refund_row["offset_strength"] = strength
        refund_row["offset_source"] = source
        refund_row["offset_rule_hint"] = rule_hint
        refund_row["offset_match_type"] = match_type
        refund_row["proposed_action"] = f"merge_refund_into:{expense_row['record_id']}"

        expense_groups.setdefault(expense_id, []).append(group)
        expense_strengths.setdefault(expense_id, []).append(strength)
        expense_sources.setdefault(expense_id, []).append(source)
        expense_rules.setdefault(expense_id, []).append(rule_hint)
        expense_match_types.setdefault(expense_id, []).append(match_type)

    for expense_id, groups in expense_groups.items():
        expense_row = by_fact_id[expense_id]
        expense_row["offset_group"] = "|".join(groups)
        expense_row["offset_role"] = "expense"
        strengths = expense_strengths.get(expense_id, [])
        expense_row["offset_strength"] = "weak" if "weak" in strengths else (strengths[0] if strengths else "")
        expense_row["offset_source"] = "|".join(expense_sources.get(expense_id, []))
        expense_row["offset_rule_hint"] = "|".join(expense_rules.get(expense_id, []))
        expense_row["offset_match_type"] = "|".join(expense_match_types.get(expense_id, []))

    return rows


def _refund_txn_base_id(txn_id: str) -> str:
    txn_id = (txn_id or "").strip()
    if not txn_id or "_" not in txn_id:
        return ""
    return txn_id.split("_", 1)[0]


def _alipay_desc_order_key(description: str) -> str:
    import re

    description = (description or "").strip()
    patterns = [
        r"美团订单-([A-Za-z0-9]+)",
        r"商户单号([A-Za-z0-9]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, description)
        if match:
            return match.group(1)
    return ""


def _refund_desc_normalized(description: str) -> str:
    description = (description or "").strip()
    prefixes = ["退款-", "退款"]
    for prefix in prefixes:
        if description.startswith(prefix):
            description = description[len(prefix):].strip()
            break
    return description


def _long_common_prefix_len(a: str, b: str) -> int:
    length = 0
    for ch_a, ch_b in zip(a, b):
        if ch_a != ch_b:
            break
        length += 1
    return length


def _refund_desc_confirms_match(ref_desc: str, exp_desc: str) -> bool:
    normalized_ref = _refund_desc_normalized(ref_desc)
    exp_desc = (exp_desc or "").strip()
    if not normalized_ref or not exp_desc:
        return False
    if normalized_ref == exp_desc:
        return True
    if normalized_ref in exp_desc or exp_desc in normalized_ref:
        return True
    return _long_common_prefix_len(normalized_ref, exp_desc) >= 16


def _wechat_device_key(description: str) -> str:
    import re

    description = (description or "").strip()
    match = re.search(r"((?:充电柜|充电插座)-[A-Za-z0-9_\-]+)", description)
    return match.group(1) if match else ""


def _wechat_meituan_order_key(description: str) -> str:
    import re

    description = (description or "").strip()
    patterns = [
        r"美团订单-(\d{20,})",
        r"-美团App-(\d{20,})",
        r"-美团微信小程序-(\d{20,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, description)
        if match:
            return match.group(1)
    return ""


def _wechat_meituan_cashier_key(description: str) -> str:
    import re

    description = (description or "").strip()
    match = re.search(r"(美团收银\d+)", description)
    return match.group(1) if match else ""


def _wechat_stable_refund_token(description: str) -> str:
    description = (description or "").strip()
    tokens = {
        "钱包充值",
        "寄件",
        "预付费充电订单",
        "自助机押金",
        "订单付款",
        "存包柜预付费",
        "先乘车后付款",
        "转账备注:微信转账",
    }
    return description if description in tokens else ""


def _wechat_refund_brand_aliases(value: str) -> set[str]:
    value = (value or "").strip()
    aliases = {value} if value else set()
    pairs = [
        ("麦当劳", "北京麦当劳食品有限公司"),
        ("UNIQLO", "优衣库"),
        ("luckin coffee", "瑞幸咖啡"),
        ("广州骑安", "滴滴"),
        ("立普世", "立普世咖啡"),
    ]
    for a, b in pairs:
        if value in {a, b}:
            aliases.update({a, b})
    return aliases


def _collect_order_based_refund_candidates(expenses: list, ref: dict, consumed: list[bool], remaining: list[float], ref_amt: float):
    matches: dict[int, str] = {}
    merchant_order_id = (ref.get("merchant_order_id", "") or "").strip()
    refund_txn_id = (ref.get("txn_id", "") or "").strip()
    desc_order_key = _alipay_desc_order_key(ref.get("note", ""))
    from ft.domain.platform_refund import alipay_order_match

    def try_add(expense_index: int, rule_hint: str):
        if consumed[expense_index]:
            return
        exp = expenses[expense_index]
        if not _refund_matches_basic_constraints(exp, ref, ref_amt, remaining[expense_index]):
            return
        matches.setdefault(expense_index, rule_hint)

    if merchant_order_id:
        for i, exp in enumerate(expenses):
            if (exp.get("merchant_order_id", "") or "").strip() == merchant_order_id:
                try_add(i, "refund_merchant_order_match")
    # FR-013: full order-key match (== / origin_ / origin*) — not rsplit-only
    if refund_txn_id:
        for i, exp in enumerate(expenses):
            origin = (exp.get("txn_id", "") or "").strip()
            if origin and alipay_order_match(refund_txn_id, origin):
                try_add(i, "import.alipay.order_prefix.v1")
    if desc_order_key:
        for i, exp in enumerate(expenses):
            exp_desc_order_key = _alipay_desc_order_key(exp.get("note", ""))
            if exp_desc_order_key and exp_desc_order_key == desc_order_key:
                try_add(i, "refund_desc_order_match")

    candidates = []
    for i, rule_hint in matches.items():
        exp = expenses[i]
        exact_amt = abs(remaining[i] - ref_amt) < Decimal("0.01")
        desc_match = bool(ref.get("note")) and (
            ref["note"] == exp.get("note", "")
            or ref["note"] in exp.get("note", "")
            or exp.get("note", "") in ref["note"]
        )
        candidates.append({
            "expense_index": i,
            "exact_amt": exact_amt,
            "desc_match": desc_match,
            "expense_date": _rec_date(exp),
            "rule_hint": rule_hint,
        })
    return candidates


def _collect_wechat_refund_candidates(expenses: list, ref: dict, consumed: list[bool], remaining: list[float], ref_amt: float):
    matches: dict[int, str] = {}
    ref_desc = ref.get("note", "")
    ref_cp = ref.get("counterparty", "")
    priority = {
        "refund_wechat_device_key": 0,
        "refund_wechat_meituan_order": 1,
        "refund_wechat_meituan_cashier": 2,
        "refund_wechat_desc_token": 3,
        "refund_wechat_brand_alias": 4,
    }

    def try_add(expense_index: int, rule_hint: str):
        if consumed[expense_index]:
            return
        exp = expenses[expense_index]
        if not _refund_matches_basic_constraints(exp, ref, ref_amt, remaining[expense_index]):
            return
        current = matches.get(expense_index)
        if current is None or priority.get(rule_hint, 99) < priority.get(current, 99):
            matches[expense_index] = rule_hint

    ref_aliases = _wechat_refund_brand_aliases(ref_cp) | _wechat_refund_brand_aliases(ref_desc)
    ref_txn_type = ref.get("txn_type", "")
    for i, exp in enumerate(expenses):
        exp_device = _wechat_device_key(exp.get("note", ""))
        exp_meituan_order = _wechat_meituan_order_key(exp.get("note", ""))
        exp_meituan_cashier = _wechat_meituan_cashier_key(exp.get("note", ""))
        exp_token = _wechat_stable_refund_token(exp.get("note", ""))
        exp_aliases = _wechat_refund_brand_aliases(exp.get("counterparty", "")) | _wechat_refund_brand_aliases(exp.get("note", ""))

        if ref_txn_type == "自助侠-退款" and exp_device:
            try_add(i, "refund_wechat_device_key")
        if "美团" in ref_txn_type and exp_meituan_order:
            try_add(i, "refund_wechat_meituan_order")
        if "美团" in ref_txn_type and exp_meituan_cashier:
            try_add(i, "refund_wechat_meituan_cashier")
        if ref_txn_type == "互联互通-退款" and exp_token == "钱包充值":
            try_add(i, "refund_wechat_desc_token")
        if ref_aliases and exp_aliases and ref_aliases & exp_aliases:
            try_add(i, "refund_wechat_brand_alias")

    candidates = []
    for i, rule_hint in matches.items():
        exp = expenses[i]
        exact_amt = abs(remaining[i] - ref_amt) < Decimal("0.01")
        candidates.append({
            "expense_index": i,
            "exact_amt": exact_amt,
            "desc_match": False,
            "expense_date": _rec_date(exp),
            "rule_hint": rule_hint,
            "rule_priority": priority.get(rule_hint, 99),
        })
    if candidates:
        best_priority = min(c["rule_priority"] for c in candidates)
        candidates = [c for c in candidates if c["rule_priority"] == best_priority]
        if any(c["exact_amt"] for c in candidates):
            candidates = [c for c in candidates if c["exact_amt"]]
    candidates.sort(key=lambda c: (-int(c["exact_amt"]), c["expense_date"]))
    for c in candidates:
        c.pop("rule_priority", None)
    return candidates


def _pair_refunds(expenses: list, refunds: list, others: list):
    """通用的退款关系识别逻辑（Alipay / WeChat / 银行卡共用）。
    返回 (fact_rows, tracking_pairs)，不在 convert 阶段净额化或删掉原始事实。
    """
    expense_ids = {id(r) for r in expenses}
    refund_ids = {id(r) for r in refunds}
    fact_rows = [r for r in others if id(r) not in expense_ids and id(r) not in refund_ids]
    fact_rows.extend(dict(exp) for exp in expenses)

    tracking_pairs = []

    for ref in sorted(refunds, key=lambda x: _rec_date(x)):
        ref_amt = abs(ref["amount"])
        candidate_expenses = expenses
        amount_budget = [abs(exp["amount"]) for exp in candidate_expenses]
        candidates = _collect_order_based_refund_candidates(candidate_expenses, ref, [False] * len(candidate_expenses), amount_budget, ref_amt)
        if not candidates and _refund_source_signal(ref) == "wechat_status":
            candidates = _collect_wechat_refund_candidates(candidate_expenses, ref, [False] * len(candidate_expenses), amount_budget, ref_amt)
        if not candidates:
            for i, exp in enumerate(candidate_expenses):
                rule_hint = "refund_cp_match"
                if not _counterparty_matches(exp["counterparty"], ref["counterparty"]):
                    raw_cp = exp.get("_raw_cp", "")
                    if not raw_cp or not _counterparty_matches(raw_cp, ref["counterparty"]):
                        continue
                    rule_hint = "refund_raw_cp_match"
                if not _refund_matches_basic_constraints(exp, ref, ref_amt, abs(exp["amount"])):
                    continue

                exact_amt = abs(abs(exp["amount"]) - ref_amt) < Decimal("0.01")
                desc_match = bool(ref["note"]) and (
                    ref["note"] == exp["note"]
                    or ref["note"] in exp["note"]
                    or exp["note"] in ref["note"]
                )
                candidates.append({
                    "expense_index": i,
                    "exact_amt": exact_amt,
                    "desc_match": desc_match,
                    "expense_date": _rec_date(exp),
                    "rule_hint": rule_hint,
                })

        if not candidates:
            for i, exp in enumerate(candidate_expenses):
                if not _refund_matches_basic_constraints(exp, ref, ref_amt, abs(exp["amount"])):
                    continue
                if not exp["note"] or not ref["note"]:
                    continue
                if (ref["note"] == exp["note"]
                        or ref["note"] in exp["note"]
                        or exp["note"] in ref["note"]):
                    exact = abs(abs(exp["amount"]) - ref_amt) < Decimal("0.01")
                    candidates.append({
                        "expense_index": i,
                        "exact_amt": exact,
                        "desc_match": True,
                        "expense_date": _rec_date(exp),
                        "rule_hint": "refund_desc_fallback",
                    })

        if not candidates:
            fact_rows.append({
                **ref,
                "counterparty": _strip_payment_prefix(ref["counterparty"]),
                "category": "income",
            })
            continue

        candidates.sort(key=lambda c: (c["exact_amt"], c["desc_match"], c["expense_date"]), reverse=True)
        best = candidates[0]
        best_idx = best["expense_index"]
        exact_amt = best["exact_amt"]
        rule_hint = best["rule_hint"]

        if ref.get("_refund_signal") == "icbc_credit_return":
            matched_expenses = [candidate_expenses[c["expense_index"]] for c in candidates]
            offset_clusters = {
                _icbc_credit_offset_cluster(exp.get("counterparty", ""), exp.get("note", ""))
                for exp in matched_expenses
            }
            account_clusters = {
                _icbc_credit_account_cluster(exp.get("payment_method", ""), exp.get("card_number", ""))
                for exp in matched_expenses
            }
            ref["_icbc_refund_same_cluster"] = len(offset_clusters) == 1 and "" not in offset_clusters
            ref["_icbc_refund_same_account_cluster"] = len(account_clusters) == 1 and "" not in account_clusters

            if ref.get("_icbc_refund_same_cluster") and ref.get("_icbc_refund_same_account_cluster"):
                amount_covering = [c for c in candidates if abs(candidate_expenses[c["expense_index"]]["amount"]) + Decimal("0.01") >= ref_amt]
                if amount_covering:
                    amount_covering.sort(key=lambda c: c["expense_date"], reverse=True)
                    best = amount_covering[0]
                    best_idx = best["expense_index"]
                    exact_amt = best["exact_amt"]
                    rule_hint = best["rule_hint"]

        if ref.get("_refund_signal") == "icbc_debit_refund":
            matched_expenses = [candidate_expenses[c["expense_index"]] for c in candidates]
            offset_clusters = {
                _icbc_debit_refund_cluster(exp.get("counterparty", ""), exp.get("note", ""))
                for exp in matched_expenses
            }
            account_clusters = {
                _icbc_debit_account_cluster(exp)
                for exp in matched_expenses
            }
            ref["_icbc_debit_refund_same_cluster"] = len(offset_clusters) == 1 and "" not in offset_clusters
            ref["_icbc_debit_refund_same_account_cluster"] = len(account_clusters) == 1 and "" not in account_clusters

            if ref.get("_icbc_debit_refund_same_cluster") and ref.get("_icbc_debit_refund_same_account_cluster"):
                amount_covering = [c for c in candidates if abs(candidate_expenses[c["expense_index"]]["amount"]) + Decimal("0.01") >= ref_amt]
                if amount_covering:
                    amount_covering.sort(key=lambda c: c["expense_date"], reverse=True)
                    best = amount_covering[0]
                    best_idx = best["expense_index"]
                    exact_amt = best["exact_amt"]
                    rule_hint = best["rule_hint"]

        if ref.get("_refund_signal") == "ccb_debit_refund":
            matched_expenses = [candidate_expenses[c["expense_index"]] for c in candidates]
            offset_clusters = {
                _ccb_refund_cluster(exp)
                for exp in matched_expenses
            }
            ref["_ccb_refund_same_cluster"] = len(offset_clusters) == 1 and "" not in offset_clusters

            if ref.get("_ccb_refund_same_cluster"):
                amount_covering = [c for c in candidates if abs(candidate_expenses[c["expense_index"]]["amount"]) + Decimal("0.01") >= ref_amt]
                if amount_covering:
                    amount_covering.sort(key=lambda c: c["expense_date"], reverse=True)
                    best = amount_covering[0]
                    best_idx = best["expense_index"]
                    exact_amt = best["exact_amt"]
                    rule_hint = best["rule_hint"]

        match_strength = _classify_refund_match(
            ref=ref,
            rule_hint=rule_hint,
            exact_amt=exact_amt,
            candidate_count=len(candidates),
            expense=candidate_expenses[best_idx],
        )

        tracking_pairs.append(_build_refund_tracking_pair(
            expense=candidate_expenses[best_idx],
            refund=ref,
            match_type="full" if exact_amt else "partial",
            rule_hint=rule_hint,
            match_strength=match_strength,
            candidate_count=len(candidates),
        ))
        fact_rows.append(dict(ref))

    fact_rows.sort(key=lambda r: ((r.get("date") or r.get("occurred_at") or ""), r.get("amount", 0)))
    return fact_rows, tracking_pairs



def _alipay_route_wealth_product(rec: dict) -> dict:
    """Route 余额宝/余利宝 as their own accounts (not destination bank / 支付宝余额).

    Product: 余额宝 and 余利宝 are book account names. Outflows from them must
    debit that account; inflows credit it. Destination bank is evidence only.
    """
    desc = str(rec.get("note") or "")
    pm = str(rec.get("payment_method") or "").strip()
    amt = rec.get("amount")
    try:
        from decimal import Decimal
        damt = Decimal(str(amt))
    except Exception:
        return rec

    # --- Out to bank: must be expense on wealth account ---
    if "转出到银行卡" in desc and "余额宝" in desc:
        rec["amount"] = -abs(damt)
        rec["category"] = "expense"
        rec["payment_method"] = "余额宝"
        # keep bank in counterparty if useful
        return rec
    if "转出到银行卡" in desc and "余利宝" in desc:
        rec["amount"] = -abs(damt)
        rec["category"] = "expense"
        rec["payment_method"] = "余利宝"
        return rec
    if "余额宝-转出到余额" in desc:
        rec["amount"] = -abs(damt)
        rec["category"] = "expense"
        rec["payment_method"] = "余额宝"
        return rec

    # --- Into 余利宝 from 支付宝余额 (source already 账户余额 expense) ---
    if "支付宝转入到余利宝" in desc:
        # keep source payment_method (账户余额) → 支付宝余额 debit; destination not on this row
        return rec

    # --- Into 余额宝 (单次转入): funding source is payment_method; destination 余额宝 needs + ---
    # Single alipay row only books source. If source is 账户余额/银行卡, leave source debit.
    # If we only have destination semantics with empty bank, force 余额宝 for 单次转入 when
    # payment already spent on source — no change to dual-entry here.
    if "余额宝-单次转入" in desc:
        # Source of funds is payment_method; do not rewrite to bank as "余额宝 account"
        # except when payment_method is already 余额宝 (reinvest).
        return rec

    # --- Income onto 余额宝 ---
    if "收益发放" in desc and "余额宝" in desc:
        rec["payment_method"] = "余额宝"
        rec["amount"] = abs(damt)
        rec["category"] = "income"
        return rec
    if "卖出至余额宝" in desc:
        rec["payment_method"] = "余额宝"
        rec["amount"] = abs(damt)
        rec["category"] = "income"
        return rec

    # Spend paid with 余额宝: payment_method already 余额宝* — mapping to 余额宝 account
    if pm.startswith("余额宝"):
        rec["payment_method"] = "余额宝"
    if pm.startswith("余利宝"):
        rec["payment_method"] = "余利宝"
    return rec


def _read_alipay_raw(path: str):
    """解析支付宝CSV，不落库，返回 list[dict]"""
    from .importers.alipay import _detect_encoding
    import csv as csv_mod

    enc = _detect_encoding(path)
    with open(path, "r", encoding=enc) as f:
        text = f.read()
    lines = text.splitlines()

    header_ln = None
    for i, line in enumerate(lines):
        if "交易时间" in line and "收/支" in line and "金额" in line:
            header_ln = i
            break
    if header_ln is None:
        print("❌ 无法找到支付宝账单表头")
        return []

    reader = csv_mod.reader(lines[header_ln:])
    header = next(reader)
    h = {col: idx for idx, col in enumerate(header)}

    raw = []  # (date_str, amount, payment_method, counterparty, description, category, txn_type)
    for row in reader:
        if len(row) < 7:
            raise ValueError(
                f"❌ 支付宝账单行缺少字段: 仅 {len(row)} 列，预期 >= 7\n"
                f"   可能是支付宝导出的格式已变更，需要更新转换器"
            )
        date_str = row[h.get("交易时间", 0)].strip()[:19].replace("/", "-")
        direction = row[h.get("收/支", 5)].strip()
        amount_str = row[h.get("金额", 6)].strip()
        txn_type = row[h.get("交易分类", 1)].strip()
        try:
            amount = Decimal(amount_str)
        except (InvalidOperation, ValueError):
            raise ValueError(
                f"❌ 支付宝账单金额无法解析: amount_str={amount_str!r}\n"
                f"   date={date_str} direction={direction} type={txn_type}"
            )
        if amount == 0:
            # 0 元交易：会员卡抵扣、积分兑换等，无资金变动，保留记录但不计入财务
            category = "expense" if direction == "支出" else "income"
        else:
            category = "expense" if amount < 0 else "income"
            if direction == "支出":
                amount = -amount
            elif direction == "收入":
                pass
            elif direction == "不计收支":
                # 支付宝提现到银行卡在原始账单里是"不计收支"，但从支付宝余额账户视角是资产流出。
                # 退款类不计收支仍保持正数，交给 _pair_refunds 配对核销。
                desc_for_direction = row[h.get("商品说明", 4)].strip() if "商品说明" in h else ""
                is_investment_outflow = (
                    txn_type == "投资理财"
                    and "收益发放" not in desc_for_direction
                    and any(k in desc_for_direction for k in ("转入", "买入", "单次转入", "支付宝转入到余利宝"))
                )
                # 余额宝/余利宝转出到银行卡：平台侧出账（负），不得记成银行卡 +income
                is_wealth_to_bank = (
                    "转出到银行卡" in desc_for_direction
                    and any(k in desc_for_direction for k in ("余额宝", "余利宝"))
                )
                is_wealth_to_balance = "余额宝-转出到余额" in desc_for_direction
                if (
                    txn_type == "账户提现"
                    or "提现-实时提现" in desc_for_direction
                    or "转出到网商银行" in desc_for_direction
                    or is_investment_outflow
                    or is_wealth_to_bank
                    or is_wealth_to_balance
                ):
                    amount = -abs(amount)
                    category = "expense"
            else:
                raise ValueError(
                    f"❌ 支付宝账单未知收/支方向: direction={direction!r}\n"
                    f"   date={date_str} type={txn_type} amount={amount_str}\n"
                    f"   请确认是否需更新转换器以支持新的方向类型"
                )

        payment_method = row[h.get("收/付款方式", 7)].strip() if "收/付款方式" in h else ""
        counterparty = row[h.get("交易对方", 2)].strip()
        desc = row[h.get("商品说明", 4)].strip() or counterparty
        txn_type = row[h.get("交易分类", 1)].strip()

        # 交易状态：用于识别"下单未付款"的假交易
        txn_status = row[h.get("交易状态", 8)].strip() if "交易状态" in h else ""

        # --- 007 whitelist skips (must comment + count; not silent) ---
        from ft.domain.platform_refund import (
            alipay_is_unpaid_closed,
            alipay_is_failed_repay,
        )
        # FR-008a: 未支付关闭 — 交易关闭/已关闭 + 非支出 + 付款方式空。
        # 下单未支付/未占用资金即关闭，导出通常无退款行；导入会造成假支出。
        if alipay_is_unpaid_closed(txn_status, direction, payment_method):
            raw.append({"_skip_reason": "unpaid_closed", "_skipped": True})
            continue
        # FR-008c: 还款失败 + 不计收支 + 付款方式空 — 自动还款未扣成；真还款见还款成功。
        if alipay_is_failed_repay(txn_status, direction, payment_method):
            raw.append({"_skip_reason": "failed_repay", "_skipped": True})
            continue
        # Paid closed expense (交易关闭|支出) and other statuses: MUST import (FR-008).
        # 0-yuan auth-hold / unfreeze / refunds: MUST import (FR-011 / FR-014a).

        if amount != 0:
            category = "expense" if amount < 0 else "income"

        normalized_cp, enriched_desc = _normalize_counterparty(counterparty, desc[:80], "alipay")
        refund_signal = _alipay_refund_signal(
            txn_type=txn_type,
            txn_status=txn_status,
            direction=direction,
            description=enriched_desc,
        )
        txn_id = row[h.get("交易订单号", 9)].strip() if "交易订单号" in h else ""
        merchant_order_id = row[h.get("商家订单号", 10)].strip() if "商家订单号" in h else ""
        fact_id = f"alipay_{txn_id}" if txn_id else f"alipay_{len(raw)+1:06d}"
        _alipay_rec = {
            "date": date_str,
            "amount": amount,
            "payment_method": payment_method,
            "counterparty": normalized_cp,
            "note": enriched_desc[:80],
            "category": category,
            "txn_type": txn_type,
            "txn_id": txn_id,
            "merchant_order_id": merchant_order_id,
            "platform_status": txn_status,
            "_alipay_direction": direction,
            "_refund_signal": refund_signal,
            "_fact_id": fact_id,
        }
        raw.append(_alipay_route_wealth_product(_alipay_rec))

    # Split whitelist skips from facts (007 acceptance counters).
    skips = [r for r in raw if r.get("_skipped")]
    facts = [r for r in raw if not r.get("_skipped")]
    for rec in facts:
        rec["platform_status"] = rec.get("platform_status") or ""
        # platform_status already set below when appending — ensure field
    # 退款配对核销（不改金额；tracking only）
    expenses = [r for r in facts if r["category"] == "expense"]
    refunds = [r for r in facts if r["amount"] > 0 and r.get("_refund_signal")]
    records, tracking_pairs = _pair_refunds(expenses, refunds, facts)
    # Stash skip stats for import acceptance (FR-002/FR-006)
    skip_counts = {"unpaid_closed": 0, "failed_repay": 0}
    for s in skips:
        reason = s.get("_skip_reason") or ""
        if reason in skip_counts:
            skip_counts[reason] += 1
    for pair in tracking_pairs:
        pair.setdefault("import_rule_id", pair.get("rule_hint") or "")
    tracking_pairs = list(tracking_pairs)
    tracking_pairs.append({
        "_acceptance": {
            "source_lines": len(raw),
            "skipped_unpaid_closed": skip_counts["unpaid_closed"],
            "skipped_failed_repay": skip_counts["failed_repay"],
            "fact_lines": len(facts),
        }
    })
    return records, tracking_pairs


def _read_wechat_raw(path: str):
    """解析微信Excel，不落库，返回 list[dict]"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要 openpyxl: pip install openpyxl")
        return []

    from .importers.wechat import INCOME_OK

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_row_i = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if row[0] and "交易时间" in str(row[0]):
            header_row_i = i
            break
    if not header_row_i:
        print("❌ 无法找到微信账单表头")
        return []

    header = [str(c or "") for c in next(ws.iter_rows(min_row=header_row_i, max_row=header_row_i, values_only=True))]
    h = {col: idx for idx, col in enumerate(header)}

    raw = []
    for row in ws.iter_rows(min_row=header_row_i + 1, values_only=True):
        if not row or not any(v for v in row if v is not None):
            continue
        vals = [str(c or "") for c in row]
        direction = vals[h["收/支"]] if "收/支" in h else ""
        status = vals[h["当前状态"]] if "当前状态" in h else ""

        # 007 FR-027: do not silently drop expense fail/closed/revoked or non-INCOME_OK income.
        # Current real-bill corpus has none of those statuses; if they appear, import them
        # (or apply FR-008b with counters when unpaid-failure can be proven). Unknown
        # neutral types still fail closed below.

        try:
            amount = Decimal(vals[h["金额(元)"]])
        except (InvalidOperation, ValueError, KeyError):
            continue
        
        # 提前提取所有字段，后面会按需使用
        payment_method = vals[h["支付方式"]] if "支付方式" in h else ""
        counterparty = vals[h["交易对方"]] if "交易对方" in h else ""
        desc = vals[h["商品"]] if "商品" in h else ""
        txn_type = vals[h["交易类型"]] if "交易类型" in h else ""
        txn_id = vals[h["交易单号"]] if "交易单号" in h else ""
        merchant_order_id = vals[h["商户单号"]] if "商户单号" in h else ""
        date_raw = vals[h["交易时间"]] if "交易时间" in h else ""
        date_str = date_raw[:19].replace("/", "-")

        if direction == "支出":
            amount = -amount
        elif direction == "收入":
            pass
        elif txn_type in ("零钱提现", "充值", "零钱充值", "零钱通存取", "理财通", "购买理财通", "信用卡还款"):
            # 微信中性交易（收/支="/"）金额列始终为正，不能按金额正负判断方向，必须按交易类型语义判断。
            if txn_type == "零钱提现":
                # 零钱 → 银行卡：本行是微信零钱出账（expense）。
                # 支付方式列是到账卡，仅作证据；mapping 用「零钱」落到微信零钱。
                # 银行卡入账由银行账单表达，再以 transfer_pair 配对。
                amount = -amount
                category = "expense"
                # Preserve destination card in counterparty if empty; force routing key.
                if payment_method and (not counterparty or counterparty in ("/", "-")):
                    counterparty = payment_method
                payment_method = "零钱"
            elif txn_type in ("充值", "零钱充值", "购买理财通", "信用卡还款"):
                # 银行卡/零钱流出到微信零钱、理财通或信用卡还款。
                amount = -amount
                category = "expense"
            elif txn_type in ("零钱通存取", "理财通"):
                text = f"{counterparty}{desc}{status}"
                if any(k in text for k in ("转出", "取出", "赎回", "到账")):
                    category = "income"
                else:
                    amount = -amount
                    category = "expense"
            else:
                continue
            # 描述为空或无意义时，用交易类型代替
            if not desc or desc in ("/", "-"):
                desc = txn_type
            normalized_cp, enriched_desc = _normalize_counterparty(counterparty, desc[:80], "wechat")
            fact_id = f"wechat_{txn_id}" if txn_id else f"wechat_{len(raw)+1:06d}"
            raw.append({
                "date": date_str,
                "amount": amount,
                "payment_method": payment_method,
                "counterparty": normalized_cp,
                "note": enriched_desc[:80],
                "category": category,
                "status": status,
                "platform_status": status,
                "txn_type": txn_type,
                "txn_id": txn_id,
                "merchant_order_id": merchant_order_id,
                "_wechat_direction": direction or "/",
                "_refund_signal": "wechat_status" if "退款" in status else "",
                "_fact_id": fact_id,
            })
            continue
        else:
            continue
        # 007: 0-yuan rows import when present (balance unaffected)
        # 描述为空或无意义时，用交易类型代替
        if not desc or desc in ("/", "-"):
            desc = txn_type

        # convert 层只看收支方向，不做语义判断
        category = "expense" if amount < 0 else "income"

        normalized_cp, enriched_desc = _normalize_counterparty(counterparty, desc[:80], "wechat")
        fact_id = f"wechat_{txn_id}" if txn_id else f"wechat_{len(raw)+1:06d}"
        raw.append({
            "date": date_str,
            "amount": amount,
            "payment_method": payment_method,
            "counterparty": normalized_cp,
            "note": enriched_desc[:80],
            "category": category,
            "status": status,
            "platform_status": status,
            "txn_type": txn_type,
            "txn_id": txn_id,
            "merchant_order_id": merchant_order_id,
            "_wechat_direction": (direction or ("支出" if category == "expense" else "收入")),
            "_refund_signal": "wechat_status" if "退款" in status else "",
            "_fact_id": fact_id,
        })

    # Dual-row refund tracking via FR-029 pure matcher; never rewrite amounts.
    from ft.domain.platform_refund import pair_wechat_refunds
    pairs = pair_wechat_refunds(raw)
    tracking_pairs = []
    for exp_i, inc_i, rule_id in pairs:
        tracking_pairs.append(_build_refund_tracking_pair(
            expense=raw[exp_i],
            refund=raw[inc_i],
            match_type="full",
            rule_hint=rule_id,
            match_strength="strong",
            candidate_count=1,
        ))
    records = list(raw)
    tracking_pairs = list(tracking_pairs)
    tracking_pairs.append({
        "_acceptance": {
            "source_lines": len(raw),
            "skipped_unpaid_closed": 0,
            "skipped_failed_repay": 0,
            "fact_lines": len(raw),
        }
    })
    return records, tracking_pairs


def _parse_icbc_lines(lines: list[str], is_credit: bool):
    """解析 ICBC PDF 文本行，返回统一 records（纯函数，方便单测）"""
    import re
    records = []

    if is_credit:
        i, current_date, current_time, current_card = 0, None, "00:00:00", ""
        while i < len(lines):
            line = lines[i].strip()
            dm = re.match(r"^(\d{4}-\d{2}-\d{2})$", line)
            if dm:
                current_date = dm.group(1)
                current_time = lines[i+1].strip() if i + 1 < len(lines) else "00:00:00"
                # 卡号在时间行下一行（16~19位纯数字）
                card_line = lines[i+2].strip() if i + 2 < len(lines) else ""
                current_card = card_line if re.match(r"^\d{16,19}$", card_line) else ""
                i += 1
                continue
            if not current_date:
                i += 1
                continue
            amt_m = re.match(r"^([+-]?[\d,]+\.\d{2})$", line)
            if amt_m:
                amount = _parse_amt(amt_m.group(1))

                # 后向扫描：找币种 + 借/贷方向（出现在金额行之前）
                currency = "CNY"
                is_charge = False
                _skip_keywords = {"交易币种", "入账币种", "入账金额", "账户余额",
                                  "对方户名", "对方账号", "摘要", "交易场所",
                                  "交易卡号", "收", "支", "消费", "活期",
                                  "钞", "汇"}
                for k in range(i - 1, max(-1, i - 20), -1):
                    if k < 0:
                        break
                    cur_raw = lines[k].strip()
                    if not cur_raw:
                        continue
                    if cur_raw == "借" and not is_charge:
                        is_charge = True
                        continue
                    if cur_raw == "贷":
                        continue
                    if cur_raw in ("美元", "USD"):
                        currency = "USD"
                        continue
                    if cur_raw in ("港币", "HKD"):
                        currency = "HKD"
                        continue
                    if cur_raw in ("日元", "JPY"):
                        currency = "JPY"
                        continue
                    if cur_raw in ("人民币",):
                        continue  # 人民币标志，继续往前找借/贷
                    if cur_raw in _skip_keywords or re.match(r"^[+-]?[\d,]+\.\d{2}$", cur_raw):
                        continue  # 已知标签或金额行，跳过
                    break  # 其他非空行→停止后向扫描
                if is_charge:
                    amount = -abs(amount)
                category = "expense" if amount < 0 else "income"

                # 从金额行向后扫描
                counterparty = ""
                description = ""
                j = i + 1
                while j < min(len(lines), i + 12):
                    s = lines[j].strip()
                    j += 1
                    if re.match(r"^[+-]?[\d,]+\.[\d]{2}$", s):
                        continue
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):  # 跳过日期行
                        continue
                    if re.match(r"^\d{2}:\d{2}:\d{2}$", s):  # 跳过时间行
                        continue
                    if "本页" in s:  # 跳过页合计行
                        continue
                    if not s:
                        continue
                    if s in ("人民币", "美元", "港币", "日元", "消费", "借", "贷",
                             "对方户名", "对方账号", "摘要", "交易场所",
                             "交易卡号", "收", "支", "交易币种", "入账币种",
                             "入账金额", "账户余额"):
                        continue
                    if s.startswith("下单时间"):  # 跳过页脚元数据行
                        continue
                    if not counterparty and not re.match(r"^\d{4,}$", s) and "****" not in s:
                        counterparty = s
                        continue
                    if counterparty and re.match(r"^\d", s) and "****" in s:
                        continue
                    if counterparty and re.match(r"^\d{4,}$", s):
                        continue
                    if s in ("转帐", "退款", "利息", "结息"):
                        continue
                    if s.startswith("手机银行") or s.startswith("网上银行"):
                        description = s
                        counterparty = counterparty if counterparty else ""
                        break
                    if any(kw in s for kw in ["美团支付", "京东支付", "财付通",
                                               "支付宝", "网银在线", "Apple",
                                               "拼多多支付", "程支付", "抖音支付"]):
                        description = s
                        break
                    if counterparty and s != counterparty:
                        if counterparty == "退货":
                            description = s  # 退货不合并，留 description 用于退款配对
                        else:
                            counterparty = counterparty + s  # 商家名换行合并
                        break
                    counterparty = s

                offset_type = _classify_icbc_credit_offset_type(counterparty, description)
                is_refund = offset_type == "merchant_refund"
                normalized_cp, enriched_desc = _normalize_counterparty(counterparty, description[:80], "icbc")
                payment_method = _infer_payment_source("icbc", normalized_cp, enriched_desc[:80])
                card_number = current_card[-4:] if current_card else ""
                fact_hash = _stable_short_hash(
                    f"{current_date} {current_time}",
                    f"{amount:.2f}",
                    currency,
                    normalized_cp,
                    enriched_desc[:80],
                    card_number,
                )
                rec = {
                    "date": f"{current_date} {current_time}",
                    "amount": amount,
                    "currency": currency,
                    "counterparty": normalized_cp,
                    "note": enriched_desc[:80],
                    "category": category,
                    "payment_method": payment_method,
                    "card_number": card_number,
                    "_raw_cp": counterparty,  # 保存原始 cp 用于退款匹配
                    "_refund_signal": "",
                    "_fact_id": f"icbc_credit_{fact_hash}",
                }
                if offset_type:
                    rec["offset_type"] = offset_type
                if is_refund:
                    merchant_text = description if counterparty == "退货" and description else counterparty
                    rec["_is_refund"] = True
                    rec["_refund_signal"] = "icbc_credit_return"
                    rec["_icbc_refund_merchant_trusted"] = not _is_icbc_credit_untrusted_merchant_text(merchant_text)
                elif offset_type in {"benefit_rebate", "campaign_cashback", "fee_reversal"}:
                    rec = _build_icbc_credit_offset_income(rec, offset_type)
                records.append(rec)
                current_date = None
            i += 1
    else:
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            amt_m = re.match(r"^([+-][\d,]+\.[\d]{2})$", line)
            if not amt_m:
                i += 1
                continue
            amount = _parse_amt(amt_m.group(1))
            date = ""
            date_line_idx = -1
            for lookback in range(1, min(11, i + 1)):
                potential = lines[i - lookback].strip()
                dm = re.match(r"^(\d{4}-\d{2}-\d{2})$", potential)
                if dm:
                    date = dm.group(1)
                    date_line_idx = i - lookback
                    break
            if not date:
                i += 1
                continue

            time_str = "00:00:00"
            if date_line_idx + 1 < len(lines):
                time_candidate = lines[date_line_idx + 1].strip()
                if re.match(r"^\d{2}:\d{2}:\d{2}$", time_candidate):
                    time_str = time_candidate

            between = [lines[j].strip() for j in range(date_line_idx + 1, i) if lines[j].strip()]
            summary = ""
            for s in between:
                if s in ("活期", "00000", "人民币", "钞", "汇", "1614", "4600", "2116", "6982"):
                    continue
                if re.match(r"^\d{2}:\d{2}:\d{2}$", s):
                    continue
                summary = s.replace("支", "").strip()
                if summary:
                    break

            lookahead = [lines[j].strip() for j in range(i + 1, min(len(lines), i + 6)) if lines[j].strip()]
            if not summary and lookahead:
                if lookahead[0] in {"退款", "退货", "撤销交易", "利息", "基金购买", "基金赎回", "银联消费", "还款"}:
                    summary = lookahead[0]

            channel = ""
            cpy = ""
            for s in lookahead:
                if s == summary:
                    continue
                if s in ("手机银行", "网上银行", "快捷支付", "其他", "批量业务", "(空)"):
                    if not channel:
                        channel = s
                    continue
                if re.match(r"^[+-]?[\d,]+\.\d{2}$", s):
                    continue
                if not cpy:
                    cpy = s

            if summary == "撤销交易":
                cpy = cpy or ""
            elif summary in {"退款", "退货"}:
                cpy = cpy or ""

            category = "expense" if amount < 0 else "income"
            normalized_cp, enriched_desc = _normalize_counterparty(cpy, summary[:80], "icbc")
            fact_hash = _stable_short_hash(
                date, time_str, f"{amount:.2f}", normalized_cp,
                enriched_desc[:80], channel,
            )
            rec = {
                "date": f"{date} {time_str}",
                "amount": amount,
                "counterparty": normalized_cp,
                "note": enriched_desc[:80] or normalized_cp[:80],
                "category": category,
                "payment_method": channel,
                "_raw_cp": cpy,
                "_refund_signal": "",
                "_fact_id": f"icbc_debit_{fact_hash}",
            }
            if summary in {"退款", "退货"} and amount > 0:
                rec["_is_refund"] = True
                rec["_refund_signal"] = "icbc_debit_refund"
                rec["counterparty"] = cpy or summary
                rec["note"] = summary
            records.append(rec)
            i += 1

    expenses = [r for r in records if r["category"] == "expense"]
    refunds = [r for r in records if r["amount"] > 0 and r.get("_is_refund")]
    if refunds:
        for r in refunds:
            if r.get("_refund_signal") == "icbc_debit_refund" and r.get("_raw_cp"):
                r["counterparty"] = r["_raw_cp"]
            elif r.get("note"):
                r["counterparty"] = r["note"]
        records, tracking_pairs = _pair_refunds(expenses, refunds, records)
        for pair in tracking_pairs:
            for key in ("expense", "refund"):
                raw = pair[key].get("counterparty", "")
                desc = pair[key].get("note", "")
                new_cp, new_desc = _normalize_counterparty(raw, desc, "icbc")
                pair[key]["counterparty"] = new_cp
                pair[key]["note"] = new_desc
        records = [r for r in records if not (
            r["category"] == "income"
            and r.get("counterparty", "") in ("消费", "财付通")
        )]
    else:
        tracking_pairs = []

    return records, tracking_pairs


def _read_icbc_raw(path: str, password: str):
    """解析工行PDF，不落库，返回 (list[dict], bill_type, tracking_pairs)"""
    from ft.importers.pdf_tools import decrypt_pdf, extract_pdf_text

    with tempfile.TemporaryDirectory(prefix="ft-icbc-") as temp_dir:
        os.chmod(temp_dir, 0o700)
        decrypted = Path(temp_dir) / "statement.pdf"
        decrypt_pdf(path, decrypted, password, timeout=30)
        text = extract_pdf_text(decrypted)

    is_credit = "信用卡" in text
    lines = text.split("\n")
    records, tracking_pairs = _parse_icbc_lines(lines, is_credit=is_credit)
    return records, "icbc_credit" if is_credit else "icbc_debit", tracking_pairs


def _pair_reversals(records: list) -> tuple[list, list]:
    """识别反向冲销交易关系，但保留双方原始事实。"""
    expenses = [(i, r) for i, r in enumerate(records) if r["category"] == "expense"]
    incomes = [(i, r) for i, r in enumerate(records)
               if r["category"] == "income" and "撤销" in r.get("note", "")]

    tracking_pairs = []
    paired_exp = set()

    for _, inc in incomes:
        inc_amt = abs(inc["amount"])
        for ei, exp in expenses:
            if ei in paired_exp:
                continue
            if abs(abs(exp["amount"]) - inc_amt) > Decimal("0.005"):
                continue
            if exp["counterparty"] != inc["counterparty"]:
                continue
            if _rec_date(exp)[:10] != _rec_date(inc)[:10]:
                continue
            if _rec_date(exp) > _rec_date(inc):
                continue

            paired_exp.add(ei)
            tracking_pairs.append({
                "expense": dict(exp),
                "refund": dict(inc),
                "match_type": "full",
                "match_strength": "strong",
                "candidate_count": 1,
                "rule_hint": "reversal_same_day_amount",
                "source_refund_signal": "reversal",
            })
            break

    return list(records), tracking_pairs


def _read_icbc_debit_raw(path: str, password: str):
    """解析工行储蓄卡PDF（表格格式），返回 (list[dict], bill_type)"""
    import pdfplumber as _pp

    pdf = _pp.open(path, password=password)
    records = []

    for page in pdf.pages:
        # 过滤水印大字（字号≥15），保留表格文字
        filtered = page.filter(lambda obj: obj.get("object_type") != "char" or obj.get("size", 99) < 15)
        tables = filtered.extract_tables()
        if not tables:
            continue
        for row in tables[0][1:]:  # 跳过表头
            if not row or not row[0]:
                continue
            rec = _parse_icbc_debit_row(row)
            if rec:
                records.append(rec)

    pdf.close()

    fact_rows = list(records)
    _, tracking_pairs = _pair_reversals(records)

    expenses = [r for r in records if r["category"] == "expense"]
    refunds = [r for r in records if r["amount"] > 0 and r.get("_is_refund")]
    others = [r for r in records if r not in expenses and r not in refunds]
    if refunds:
        _, refund_pairs = _pair_refunds(expenses, refunds, others)
        tracking_pairs.extend(refund_pairs)
    return fact_rows, "icbc_debit", tracking_pairs


def _parse_icbc_debit_row(row: list) -> dict | None:
    """解析储蓄卡PDF的一行表格数据"""
    import re as _re

    dt_str = (row[0] or "").replace("\n", " ")
    dm = _re.search(r"(\d{4}-\d{2}-\d{2})", dt_str)
    tm = _re.search(r"(\d{2}:\d{2}:\d{2})", dt_str)
    if not dm:
        raise ValueError(
            "❌ 工行借记卡行无法提取日期\n"
            f"   可能是PDF格式变更，需要更新转换器"
        )
    # 保护：pdfplumber 可能返回截断行
    if len(row) < 13:
        raise ValueError(
            f"❌ 工行借记卡行列数不足: 仅 {len(row)} 列，预期 >= 13\n"
            f"   可能是pdfplumber解析结果截断或PDF格式变更"
        )
    date = f"{dm.group(1)} {tm.group(1) if tm else '00:00:00'}"

    # 币种
    cur_str = (row[4] or "").replace("\n", "").strip()
    cur_map = {"人民币": "CNY", "美元": "USD", "港币": "HKD", "日元": "JPY"}
    currency = cur_map.get(cur_str, "CNY")

    # 金额 — 水印已过滤，干净的数字
    amt_str = (row[8] or "").replace("\n", "").strip()
    amt_m = _re.search(r"([+-])?([\d,]+\.[\d]{2})", amt_str)
    if not amt_m:
        raise ValueError(
            "❌ 工行借记卡行金额无法解析\n"
            f"   可能是PDF格式变更或水印干扰"
        )
    sign = amt_m.group(1) or ""
    num = Decimal(amt_m.group(2).replace(",", ""))
    amount = -num if sign == "-" else num

    # 摘要 — 匹配已知关键词（水印噪声可能把关键词拆散）
    summary = _match_debit_summary((row[6] or "").replace("\n", "").strip())

    # 对方户名
    counterparty = (row[10] or "").replace("\n", "").strip()

    # 清洗对方户名：摘要文本可能因水印/换行混入 counterparty
    if counterparty and summary:
        # 用 _is_subseq 检测（乱码后字符不一定连续出现）
        if _is_subseq("基金快速赎回", counterparty) or _is_subseq("基金赎回", counterparty):
            counterparty = "中国工商银行股份有限公司基金清算专户"
        elif _is_subseq("基金购买", counterparty):
            counterparty = "中国工商银行股份有限公司基金清算专户"

    # 渠道
    channel = (row[12] or "").replace("\n", "").strip()

    category = "expense" if amount < 0 else "income"

    debit_offset_type = ""
    if summary == "撤销交易":
        debit_offset_type = "reversal"
    elif summary in {"退款", "退货"}:
        debit_offset_type = "refund"

    fact_hash = _stable_short_hash(date, f"{amount:.2f}", counterparty, summary, channel)
    return {
        "date": date,
        "amount": amount,
        "currency": currency,
        "counterparty": counterparty,
        "note": summary,
        "category": category,
        "payment_method": channel,
        "_raw_cp": counterparty,
        "_fact_id": f"icbc_debit_{fact_hash}",
        "_debit_offset_type": debit_offset_type,
        "_is_refund": debit_offset_type == "refund",
        "_is_reversal": debit_offset_type == "reversal",
        "_refund_signal": "icbc_debit_refund" if debit_offset_type == "refund" and amount > 0 else "",
    }


def _match_debit_summary(text: str) -> str:
    """从含噪声的摘要文本中匹配已知交易类型关键词"""
    import re as _re

    # 已知摘要关键词，按频率/重要性排序
    KWS = [
        "支付宝转账", "无卡支付", "工资", "转账", "购汇还款", "撤销交易",
        "个人购汇", "跨境汇款", "利息", "基金购买", "基金赎回",
        "他行汇入", "银联入账", "预约购汇", "网转", "跨行汇款",
    ]

    # 去除非中文噪声，保留中文字符
    clean = _re.sub(r"[^\u4e00-\u9fff]", "", text)

    for kw in KWS:
        # 检查关键词中的所有字是否都在clean中按顺序出现
        if _is_subseq(kw, clean):
            return kw
    # 兜底：回退原始文本（去噪声）
    return clean


def _is_subseq(pattern: str, text: str) -> bool:
    """检查 pattern 的字符是否按顺序出现在 text 中（用于噪声中匹配关键词）"""
    it = iter(text)
    return all(ch in it for ch in pattern)


def _parse_amt(s: str) -> Decimal:
    s = s.strip().replace(",", "").replace("+", "")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _build_output_row(
    rec: dict, *, bill_type: str, account: str | None = None,
    currency: str | None = None, rules=None, default_action: str = "error",
) -> dict:
    """Build a normalized output row; route account via mapping unless explicitly given.

    Prefer bill_type + card_number composite source, then bill_type + payment_method.
    Longer mapping match wins (see ft.mapping.match_payment_method).
    """
    from .mapping import match_payment_method

    if account:
        acct_name = account
        cur = currency or "CNY"
    else:
        if rules is None:
            from .mapping import load_rules
            rules, default_action = load_rules()
        card_num = rec.get("card_number", "") or ""
        match = None
        if card_num:
            match = match_payment_method(rules, f"{bill_type}_{card_num}", "*")
        if not match:
            match = match_payment_method(
                rules, bill_type, rec.get("payment_method", "") or ""
            )
        if match:
            acct_name = match["account"]
            cur = match.get("currency") or currency or "CNY"
        else:
            action = (default_action or "error").lower()
            detail = (
                f"source={bill_type} payment_method='{rec.get('payment_method', '')}' "
                f"card_number='{card_num}' counterparty='{rec.get('counterparty', '')}' "
                f"amount={rec.get('amount', '')}"
            )
            if action in {"error", "fail"}:
                raise ValueError(
                    f"未匹配 mapping 规则: {detail}\n"
                    f"  请在 ~/.ft/mapping.yaml 中添加映射规则后重试"
                )
            if action == "skip":
                return None  # type: ignore[return-value]
            raise ValueError(
                f"未匹配 mapping 规则且 default='{default_action}': {detail}"
            )

    payment_src = _infer_payment_source(
        bill_type,
        rec.get("counterparty", ""),
        rec.get("note", ""),
    )

    cpy = rec.get("counterparty", "")
    if bill_type == "icbc_credit" or bill_type == "icbc_debit":
        cpy = _strip_payment_prefix(cpy)

    # Only provider-owned identifiers are safe for overlap idempotency.  The
    # parser's `_fact_id` fallback may be row-position based; leaving it empty
    # lets the import service derive a canonical content identity instead.
    provider_record_id = ""
    if bill_type in {"alipay", "wechat"}:
        provider_record_id = rec.get("txn_id") or rec.get("merchant_order_id") or ""
    elif bill_type in {"icbc_credit", "icbc_debit", "ccb_debit"}:
        provider_record_id = rec.get("_fact_id", "")

    row_currency = rec.get("currency", cur) or cur
    if currency and not rec.get("currency"):
        row_currency = currency

    return {
        "record_id": provider_record_id,
        "date": _rec_date(rec),
        "amount": rec["amount"],
        "currency": str(row_currency).upper(),
        "counterparty": cpy,
        "note": rec.get("note", ""),
        "category": rec["category"],
        "account_name": acct_name,
        "source": payment_src,
        "bill_source": bill_type,
        "transfer_account": rec.get("transfer_account", ""),
        "locked": rec.get("locked", ""),
        "offset_group": rec.get("offset_group", ""),
        "offset_role": rec.get("offset_role", ""),
        "offset_strength": rec.get("offset_strength", ""),
        "offset_source": rec.get("offset_source", ""),
        "offset_rule_hint": rec.get("offset_rule_hint", ""),
        "offset_match_type": rec.get("offset_match_type", ""),
        "proposed_action": rec.get("proposed_action", "leave_as_is"),
    }


def _prepare_convert_rows(path: str, source: str, password: str = None):
    tracking_pairs = []
    if source == "icbc":
        rows, bill_type, tracking_pairs = _read_icbc_raw(path, password)
        if not rows:
            return [], "", []
    elif source == "icbc-debit":
        rows, bill_type, tracking_pairs = _read_icbc_debit_raw(path, password)
        if not rows:
            return [], "", []
    elif source == "alipay":
        rows, tracking_pairs = _read_alipay_raw(path)
        bill_type = "alipay"
    elif source == "wechat":
        rows, tracking_pairs = _read_wechat_raw(path)
        bill_type = "wechat"
    elif source == "ccb-debit":
        from .importers.ccb_debit import read_ccb_debit
        rows, _ = read_ccb_debit(path)
        bill_type = "ccb_debit"
        expenses = [r for r in rows if r["category"] == "expense"]
        refunds = []
        others = []
        for row in rows:
            if row["category"] == "income" and row.get("_ccb_refund_signal"):
                refunds.append({**row, "_refund_signal": row["_ccb_refund_signal"]})
            elif row["category"] == "expense":
                continue
            else:
                others.append(row)
        rows, tracking_pairs = _pair_refunds(expenses, refunds, others)
    else:
        print(f"❌ 未知账单类型: {source}")
        return [], "", []

    # Pull 007 acceptance meta rows out before fact building
    # Acceptance already appended onto tracking_pairs by readers when available.
    tracking_pairs = list(tracking_pairs or [])
    rows = _build_convert_fact_rows(rows, [p for p in tracking_pairs if not p.get("_acceptance")])
    rows = _attach_tracking_metadata(rows, [p for p in tracking_pairs if not p.get("_acceptance")])
    return rows, bill_type, tracking_pairs
