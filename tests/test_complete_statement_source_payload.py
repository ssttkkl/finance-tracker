"""现金账单来源行快照和对方账号的回归测试。"""
from __future__ import annotations

import csv
from pathlib import Path

import pytest
from sqlalchemy import select


def _write_mapping(path: Path) -> None:
    import yaml

    path.write_text(
        yaml.safe_dump(
            {
                "rules": [{
                    "source": "alipay",
                    "match": "账户余额",
                    "account": "支付宝余额",
                    "currency": "CNY",
                }],
                "default": "error",
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )


def test_alipay_source_payload_is_complete_raw_row_and_persists_counterparty_account(
    tmp_path, monkeypatch,
):
    from ft import mapping as mapping_mod
    from ft.adapters.relational.models import CashTransactionModel
    from ft.adapters.statement_import import StatementParser
    from ft.application.statement_import import StatementImportService
    from ft.domain.imports import StatementImportCommand
    from test_postgres_adapter import _database

    header = [
        "交易时间", "交易分类", "交易对方", "对方账号", "商品说明", "收/支", "金额",
        "收/付款方式", "交易状态", "交易订单号", "商家订单号", "备注",
    ]
    source_row = [
        "2026-08-03 10:00:00", "账户提现", "示例银行", "demo***@example.com", "提现-实时提现",
        "不计收支", "88.00", "账户余额", "交易成功", "PAYLOAD-001", "", "",
    ]
    source = tmp_path / "alipay.csv"
    with source.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerow(source_row)

    mapping = tmp_path / "mapping.yaml"
    _write_mapping(mapping)
    monkeypatch.setattr(mapping_mod, "MAPPING_PATH", mapping)
    sessions, unit_of_work = _database()
    with unit_of_work(sessions, "workspace-a") as uow:
        uow.accounts.add_raw({"name": "支付宝余额", "type": "cash"})
        uow.commit()

    result = StatementImportService(
        unit_of_work(sessions, "workspace-a"), StatementParser(),
    ).import_statement(StatementImportCommand(source_path=str(source), source="alipay"))

    assert result.count == 1
    with sessions() as session:
        fact = session.scalar(select(CashTransactionModel))
        assert fact is not None
        assert fact.counterparty_account == "demo***@example.com"
        assert fact.counterparty_account_attrs == ["masked"]
        assert fact.source_payload == dict(zip(header, source_row, strict=True))
        assert not {"account_name", "record_type", "source_type", "payment_method"} & set(
            fact.source_payload
        )


def test_alipay_rejects_rows_that_cannot_preserve_every_source_column(tmp_path):
    from ft.convert import _read_alipay_raw

    header = ["交易时间", "交易分类", "交易对方", "商品说明", "收/支", "金额", "收/付款方式"]
    source = tmp_path / "alipay-incomplete.csv"
    source.write_text(
        ",".join(header) + "\n" + ",".join(["2026-08-03", "消费", "示例商户", "商品", "支出", "12.34"]) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="字段数量不匹配"):
        _read_alipay_raw(str(source))


def test_alipay_source_payload_preserves_quoted_newline_values(tmp_path):
    from ft.convert import _read_alipay_raw

    header = ["交易时间", "交易分类", "交易对方", "商品说明", "收/支", "金额", "收/付款方式", "备注"]
    source_row = [
        "2026-08-03 10:00:00", "消费", "示例商户", "示例商品", "支出", "12.34", "账户余额", "第一行\n第二行",
    ]
    source = tmp_path / "alipay-multiline.csv"
    with source.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerow(source_row)

    records, _tracking = _read_alipay_raw(str(source))

    assert records[0]["_source_payload"] == dict(zip(header, source_row, strict=True))


def test_alipay_source_payload_preserves_single_unnamed_source_column(tmp_path):
    from ft.convert import _read_alipay_raw

    header = ["交易时间", "交易分类", "交易对方", "商品说明", "收/支", "金额", "收/付款方式", ""]
    source_row = [
        "2026-08-03 10:00:00", "消费", "示例商户", "示例商品", "支出", "12.34", "账户余额", "",
    ]
    source = tmp_path / "alipay-unnamed-column.csv"
    with source.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerow(source_row)

    records, _tracking = _read_alipay_raw(str(source))

    assert records[0]["_source_payload"] == dict(zip(header, source_row, strict=True))


def test_wechat_withdrawal_keeps_raw_payment_method_and_extracts_destination_card(tmp_path):
    from openpyxl import Workbook

    from ft.convert import _read_wechat_raw

    header = [
        "交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)", "支付方式",
        "当前状态", "交易单号", "商户单号",
    ]
    source_row = [
        "2026-08-03 10:00:00", "零钱提现", "/", "/", "/", "12.34", "示例银行储蓄卡(4321)",
        "已到账", "WX-001", "",
    ]
    source = tmp_path / "wechat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["微信支付账单"])
    sheet.append(header)
    sheet.append(source_row)
    workbook.save(source)

    records, _tracking = _read_wechat_raw(str(source))

    assert len(records) == 1
    assert records[0]["counterparty_account"] == "示例银行储蓄卡(4321)"
    assert records[0]["_source_payload"] == dict(zip(header, source_row, strict=True))

    from ft.convert import _build_output_row

    output = _build_output_row(records[0], bill_type="wechat", account="微信零钱")
    assert output["counterparty_account"] == "4321"
    assert output["counterparty_account_attrs"] == ["tail"]


def test_wechat_source_without_a_counterparty_account_keeps_the_field_empty(tmp_path):
    from openpyxl import Workbook

    from ft.convert import _read_wechat_raw

    header = [
        "交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)", "支付方式",
        "当前状态", "交易单号", "商户单号",
    ]
    source_row = [
        "2026-08-03 10:00:00", "商户消费", "示例商户", "示例商品", "支出", "12.34", "零钱",
        "支付成功", "WX-002", "",
    ]
    source = tmp_path / "wechat-no-counterparty-account.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["微信支付账单"])
    sheet.append(header)
    sheet.append(source_row)
    workbook.save(source)

    records, _tracking = _read_wechat_raw(str(source))

    assert len(records) == 1
    assert records[0]["counterparty_account"] == ""
    assert records[0]["_source_payload"] == dict(zip(header, source_row, strict=True))


def test_ccb_keeps_full_row_and_extracts_only_counterparty_account(tmp_path):
    import xlwt

    from ft.importers.ccb_debit import read_ccb_debit

    header = ["序号", "摘要", "币别", "保留列", "交易日期", "交易金额", "余额", "交易地点", "对方账号与户名"]
    source_row = [1, "转账支取", "人民币元", "", 20260803, -50.25, 100.00, "***", "6222****4321/示例户名"]
    source = tmp_path / "ccb.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("明细")
    sheet.write(0, 0, "建行明细")
    sheet.write(1, 0, "卡号/账号:6222000012345678")
    sheet.write(2, 0, "汇总")
    for index, value in enumerate(header):
        sheet.write(3, index, value)
    for index, value in enumerate(source_row):
        sheet.write(4, index, value)
    workbook.save(str(source))

    records, _tracking = read_ccb_debit(str(source))

    assert len(records) == 1
    assert records[0]["counterparty_account"] == "6222****4321"
    assert records[0]["_source_payload"] == {
        key: str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
        for key, value in zip(header, source_row, strict=True)
    }

    from ft.convert import _build_output_row

    output = _build_output_row(records[0], bill_type="ccb_debit", account="建行储蓄卡")
    assert output["counterparty_account"] == "6222****4321"
    assert output["counterparty_account_attrs"] == ["masked"]


def test_icbc_debit_keeps_full_table_row_and_extracts_counterparty_account():
    from ft.convert import _parse_icbc_debit_row

    header = [
        "交易日期", "交易时间", "本方账号", "流水号", "币种", "发生地", "摘要", "余额", "发生额",
        "对方行名", "对方户名", "对方账号", "交易渠道",
    ]
    source_row = [
        "2026-08-03", "10:00:00", "6222****0000", "ICBC-001", "人民币", "北京", "转账", "100.00",
        "-88.00", "示例银行", "示例户名", "6222****4321", "手机银行",
    ]
    source_payload = dict(zip(header, source_row, strict=True))

    record = _parse_icbc_debit_row(source_row, source_payload=source_payload)

    assert record is not None
    assert record["counterparty_account"] == "6222****4321"
    assert record["_source_payload"] == source_payload

    from ft.convert import _build_output_row

    output = _build_output_row(record, bill_type="icbc_debit", account="工行借记卡")
    assert output["counterparty_account"] == "6222****4321"
    assert output["counterparty_account_attrs"] == ["masked"]


def test_icbc_debit_uses_pdf_account_column_as_source_account_identity():
    from ft.convert import _parse_icbc_debit_row

    account = "1614020101021984636"
    row = [
        "2026-08-03\n10:00:00", account, "活期", "00001", "人民币", "钞", "消费", "北京",
        "-88.00", "100.00", "示例户名", "6222****4321", "快捷支付",
    ]

    record = _parse_icbc_debit_row(row)

    assert record is not None
    assert record["_source_account_identifier"] == account
    assert record["file_account_key"] == account
    assert record["source_display_name"] == "工商银行借记卡"


def test_icbc_credit_transfer_extracts_structured_masked_counterparty_account():
    from ft.convert import _build_output_row, _parse_icbc_lines

    lines = [
        "2026-08-03", "10:00:00", "6222000000000000", "借", "人民币", "88.00",
        "人民币", "88.00", "100.00", "转帐", "6222****4321", "",
    ]

    records, _tracking = _parse_icbc_lines(lines, is_credit=True)

    assert len(records) == 1
    assert records[0]["counterparty_account"] == "6222****4321"
    output = _build_output_row(records[0], bill_type="icbc_credit", account="工行信用卡")
    assert output["counterparty_account"] == "6222****4321"
    assert output["counterparty_account_attrs"] == ["masked"]


def test_icbc_credit_transfer_does_not_choose_between_multiple_masked_accounts():
    from ft.convert import _parse_icbc_lines

    lines = [
        "2026-08-03", "10:00:00", "6222000000000000", "借", "人民币", "88.00",
        "转帐", "6222****4321", "9558****7654", "",
    ]

    records, _tracking = _parse_icbc_lines(lines, is_credit=True)

    assert len(records) == 1
    assert records[0]["counterparty_account"] == ""


def test_icbc_asia_preserves_non_numeric_counterparty_account_identifier():
    from ft.convert import _build_output_row

    record = {
        "date": "2026-08-03 10:00:00",
        "amount": "12.34",
        "currency": "HKD",
        "counterparty": "示例对方",
        "counterparty_account": "counterparty@example.com",
        "note": "本地转账",
        "category": "income",
        "payment_method": "工银亚洲活期账户",
        "txn_type": "轉賬",
        "summary": "本地轉賬",
        "_fact_id": "fixture",
    }

    output = _build_output_row(record, bill_type="icbc_asia", account="工银亚洲账户")

    assert output["counterparty_account"] == "counterparty@example.com"
    assert output["counterparty_account_attrs"] == ["full"]
